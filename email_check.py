# -*- coding: utf-8 -*-
"""
@Tool : PyCharm

@User : 21407

@File : email_check.py

@Email: yypqcaa@163.com

@Date : 2023/2/27 22:31

@Desc :
"""

import os
import time
import socket
import dns.resolver
from loguru import logger
from openpyxl import Workbook


def __re_check(s_email: str) -> bool:
    """检验邮箱格式
    :param s_email: 邮箱地址
    :return: True|False
    """
    is_legal = False
    try:
        is_legal = isinstance(s_email, str) and s_email.count('@') == 1
    except Exception as e:
        logger.error(f"{e.__traceback__.tb_lineno}:--:{e}")
    return is_legal


def __do_dns(s_email) -> str:
    str_ret = ''
    try:
        my_resolver = dns.resolver.Resolver()
        my_resolver.nameservers = ['223.5.5.5', '8.8.8.8', '192.168.2.1']
        obj_mx = my_resolver.resolve(s_email, dns.rdatatype.MX)
        # 可能有多条 找到优先级最大的
        int_max = 0
        for obj_i in obj_mx.response.answer:
            for j in obj_i.items:
                curr_max, str_mx = j.preference, '.'.join([i.decode('utf-8') for i in j.exchange.labels if i])
                if curr_max >= int_max:
                    int_max = curr_max
                    str_ret = str_mx
    except Exception as e:
        logger.error(f"{e.__traceback__.tb_lineno}:--:{e}")
    return str_ret


def check_email(lst_email: list, obj):
    lst_ret = []
    try:
        for str_email in lst_email[:100]:
            try:
                int_ratio = 0
                lst_step = []
                int_port = 25  # smpt 端口
                int_timeout = 20  # 超时时间
                str_mx = ''  # 解析邮件服务器地址
                client_session = None
                # 校验邮箱格式
                is_legal = __re_check(str_email)
                int_ratio = 5 if is_legal else int_ratio
                obj.show_message('', '', f"{str_email}_{int_ratio}: {is_legal}")
                lst_step.append({'stepId': 1, 'stepName': "步骤[ 1 ]：Email格式校验", 'isOk': True,
                                 'isOkType': 'success' if is_legal else 'error',
                                 'stepDate': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())),
                                 'stepRemark': "成功，邮箱格式符合规范" if is_legal else '失败，邮箱格式不符合规范'
                                 })
                # 解析 dns
                if int_ratio == 5:
                    str_mx = __do_dns(str_email.split('@')[-1])
                    obj.show_message('', '', f"{str_email}_{int_ratio}: {str_mx}")
                    int_ratio = 20 if str_mx else int_ratio
                    lst_step.append({'stepId': 2, 'stepName': "步骤[ 2 ]：获取邮件服务器地址", 'isOk': True,
                                     'isOkType': 'success' if str_mx else 'error',
                                     'stepDate': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())),
                                     'stepRemark': f"成功，邮件服务器地址：{str_mx}" if str_mx else '获取邮件服务器地址失败'
                                     })
                if int_ratio == 20 and str_mx:
                    str_tele = ''
                    try:
                        # connect 超时时间
                        socket.setdefaulttimeout(int_timeout)
                        client_session = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                        client_session.connect((str_mx, int_port))
                    except Exception as e:
                        obj.show_message('', '', f"{e.__traceback__.tb_lineno}: {e}")
                    else:
                        str_tele = client_session.recv(1024).decode('utf-8')
                        if str_tele.startswith('2'):
                            int_ratio = 35
                        obj.show_message('', '', f"{str_email}_{int_ratio}: {str_tele}")
                    finally:
                        lst_step.append({'stepId': 3, 'stepName': "步骤[ 3 ]：与服务器建立连接", 'isOk': True if client_session else False,
                                         'isOkType': 'success' if int_ratio == 35 else 'error',
                                         'stepDate': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())),
                                         'stepRemark': f"{'成功' if int_ratio == 35 else '失败'}，服务器返回： {str_tele}"
                                         })
                # 与服务器见了通信
                if int_ratio == 35 and client_session:
                    str_tele = ''
                    try:
                        client_session.send(b'HELO HELLO\n')
                        str_tele = client_session.recv(1024).decode('utf-8')
                        if str_tele.startswith('2'):
                            int_ratio = 50
                        obj.show_message('', '', f"{str_email}_{int_ratio}: {str_tele}")
                    except Exception as e:
                        obj.show_message('', '', f"{e.__traceback__.tb_lineno}: {e}")
                    finally:
                        lst_step.append({'stepId': 4, 'stepName': "步骤[ 4 ]：与服务器建立通信", 'isOk': True,
                                         'isOkType': 'success' if int_ratio == 50 else 'error',
                                         'stepDate': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())),
                                         'stepRemark': f"{'成功' if int_ratio == 50 else '失败'}，服务器返回： {str_tele}"
                                         })
                # 模拟发送邮件
                if int_ratio == 50 and client_session:
                    str_tele = ''
                    try:
                        client_session.send('mail from:<23648646@qq.com>\n'.encode('utf-8'))
                        str_tele = client_session.recv(1024).decode('utf-8')
                        if not str_tele.startswith('5'):
                            int_ratio = 65
                        obj.show_message('', '', f"{str_email}_{int_ratio}: {str_tele}")
                    except Exception as e:
                        obj.show_message('', '', f"{e.__traceback__.tb_lineno}: {e}")
                    finally:
                        lst_step.append({'stepId': 5, 'stepName': "步骤[ 5 ]：模拟发送测试邮件", 'isOk': True,
                                         'isOkType': 'success' if int_ratio == 65 else 'error',
                                         'stepDate': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())),
                                         'stepRemark': f"{'成功' if int_ratio == 65 else '失败'}，服务器返回： {str_tele}"
                                         })
                # 核对联系人信息
                if int_ratio == 65:
                    str_tele = ''
                    try:
                        client_session.send(f'RCPT TO:<{str_email}>\n'.encode('utf-8'))
                        str_tele = client_session.recv(1024).decode('utf-8')
                        if not str_tele.startswith('5'):
                            int_ratio = 100
                        obj.show_message('', '', f"{str_email}_{int_ratio}: {str_tele}")
                    except Exception as e:
                        obj.show_message('', '', f"{e.__traceback__.tb_lineno}: {e}")
                    finally:
                        lst_step.append({'stepId': 6, 'stepName': "步骤[ 6 ]：核对联系人邮箱地址", 'isOk': True,
                                         'isOkType': 'success' if int_ratio == 100 else 'error',
                                         'stepDate': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())),
                                         'stepRemark': f"{'成功' if int_ratio == 100 else '失败'}，服务器返回： {str_tele}"
                                         })
                obj.show_message('', '', f'email: {str_email} 解析完成...')
                lst_ret.append({
                    'email': str_email,
                    'isOk': True,
                    'isOkType': 'success' if int_ratio == 100 else 'error',
                    'ratio': int_ratio,
                    'stepList': lst_step,
                    'str_time': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                })
            except Exception as err:
                obj.show_message('', '', f"{str_email}: {err.__traceback__.tb_lineno}: {err}")
                logger.error(f"{err.__traceback__.tb_lineno}: {err}")
        else:
            obj.show_message('', '', f'所有邮箱解析完成,开始写入excel中....')
            __write_excel(lst_ret, obj)
    except Exception as err:
        logger.error(f"{err.__traceback__.tb_lineno}:--:{err}")


def __write_excel(lst_info: list, obj):
    str_path = os.path.join(os.path.abspath('.'), 'email_check')
    str_file = os.path.join(str_path, f"check_{time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))}.xlsx")
    lst_title = ['邮箱地址', '验证日期', '第一步', '第二步', '第三步', '第四步', '第五步', '第六步', '结果', '有效性概率(%)']
    if not os.path.isdir(str_path):
        os.makedirs(str_path, exist_ok=True)
    try:
        wb = Workbook()
        # 默认工作簿
        ws_succ = wb.active
        ws_succ.title = '校验成功'
        # 创建校验失败的工作簿
        ws_fail = wb.create_sheet('校验失败', 1)
        # 写入表头
        for int_index, str_title in enumerate(lst_title, 1):
            ws_succ.cell(1, int_index).value = str_title
            ws_fail.cell(1, int_index).value = str_title
        int_succ = 2
        int_fail = 2
        for dit_info in lst_info:
            if dit_info['isOkType'] == 'success':
                obj_sheet = ws_succ
                int_index = int_succ
                int_succ += 1
            else:
                obj_sheet = ws_fail
                int_index = int_fail
                int_fail += 1
            obj_sheet.cell(int_index, 1).value = dit_info['email']
            obj_sheet.cell(int_index, 2).value = dit_info['str_time']
            for int_i, dit_step in enumerate(dit_info['stepList'], 3):
                obj_sheet.cell(int_index, int_i).value = dit_step['stepRemark']
            obj_sheet.cell(int_index, 9).value = '成功' if dit_info['isOkType'] == 'success' else '失败'
            obj_sheet.cell(int_index, 10).value = dit_info['ratio']
        wb.save(f'{str_file}')
        wb.close()
        obj.show_message('', '', f'邮箱解析结果写入excel成功,保存路径: {str_file}')
    except Exception as e:
        obj.show_message('', '', f"写入excel失败, {e.__traceback__.tb_lineno}: {e}")
        logger.error(f"{e.__traceback__.tb_lineno}: {e}")
