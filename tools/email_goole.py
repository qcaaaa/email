# -*- coding: utf-8 -*-
"""
@Tool : PyCharm

@User : 21407

@File : email_check.py

@Email: yypqcaa@163.com

@Date : 2023/3/13 22:31

@Desc :
"""

import os
import time
import pyautogui
from itertools import product
from openpyxl import Workbook
from loguru import logger
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from constant import DRIVER_PATH

int_timeout = 60


def search(city, keyword, self_ui):
    lst_data = []
    lst_search = list(product(city.split('\n'), keyword.split('\n')))
    self_ui.show_message('', '', f'本轮一共 {len(lst_search)} 种搜索组合')
    for tuple_search in lst_search:
        driver = None
        try:
            # # # 创建ChromeOptions实例
            # options = Options()
            # # # 只加载 html
            # options.page_load_strategy = "eager"
            # 创建 Chrome 实例
            # driver = webdriver.Chrome(service=Service(DRIVER_PATH), options=options)
            driver = webdriver.Chrome(service=Service(DRIVER_PATH))
            # 全屏
            driver.maximize_window()
            # 打开谷歌地图
            driver.get('https://www.google.com/maps')
        except Exception as err_:
            logger.error(f'打开谷歌地图失败: {err_.__traceback__.tb_lineno}: {err_}')
        else:
            self_ui.show_message('', '', f'成功打开谷歌地图')
            str_city, str_key = tuple_search
            try:
                # 在搜索栏输入关键字
                # 等待页面加载完成
                search_box = WebDriverWait(driver, int_timeout).until(
                    EC.presence_of_element_located((By.ID, "searchboxinput"))
                )
                search_box.clear()
                search_box.send_keys(str_city)
                search_box.send_keys(Keys.RETURN)
                self_ui.show_message('', '', f'成功定位至 {str_city}')
                time.sleep(5)
                # 点击附近按钮
                button = WebDriverWait(driver, int_timeout).until(
                    EC.presence_of_element_located((By.XPATH, "//div[contains(text(), '附近')]"))
                )
                button.click()
                time.sleep(5)
                # 在搜索栏输入关键字
                search_box = WebDriverWait(driver, int_timeout).until(
                    EC.presence_of_element_located((By.ID, 'searchboxinput'))
                )
                search_box.send_keys(str_key)
                search_box.send_keys(Keys.RETURN)
                self_ui.show_message('', '', f'开始搜索关键字: {str_key}')
                time.sleep(5)

                # 滑动 获取全部搜索结果
                if __load(driver, self_ui, str_key):

                    # 获取所有搜索的 超链接
                    a_tags = WebDriverWait(driver, int_timeout).until(
                        EC.presence_of_all_elements_located((By.CLASS_NAME, 'hfpxzc'))
                    )
                    self_ui.show_message('', '', f'开始解析数据')
                    lst_data.extend(__check_url(a_tags, driver, self_ui, str_city, str_key))
            except Exception as err:
                logger.error(f'定位失败: {err.__traceback__.tb_lineno}: {err}')
        finally:
            # 关闭浏览器
            if driver:
                driver.quit()
    else:
        self_ui.show_message('', '', f'全部搜索完毕, 共搜索出 {len(lst_data)} 组信息')
        if lst_data:
            self_ui.show_message("", "", f"开始准备写入")
            __write_excel(lst_data, self_ui)
        self_ui.obj_ui.google_button.setEnabled(True)
    return


def __load(driver, ui, str_key):
    element = None
    try:
        tuple_size = pyautogui.size()
        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, f'// *[ @ aria-label="与“{str_key}”相符的搜索结果"]'))
            )
        except:
            ui.show_message('', '', f'无法通过页面定位, 开始控制鼠标移动, 请勿移动鼠标...')
            pyautogui.moveTo(70, tuple_size[1] // 2)

        s_time = time.time()

        while time.time() - s_time <= 10 * 60:
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "HlvSq"))
                )
            except:
                if not element:
                    pyautogui.scroll(-(tuple_size[0] // 2))
                else:
                    driver.execute_script('arguments[0].style.overflow = "auto";arguments[0].scrollTop = 100000', element)
                ui.show_message('', '', f'搜索更多结果中....')
            else:
                ui.show_message('', '', f'所有结果搜索完毕')
                break
    except Exception as err:
        logger.error(f'{err.__traceback__.tb_lineno}: {err}')
        return False
    return True


def __get(driver, ui, sty_type, str_path, str_key) -> str:
    str_info = ''
    try:
        element = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((sty_type, str_path))
        )
        str_info = element.get_attribute(str_key)
    except Exception as err1:
        logger.error(f'{err1.__traceback__.tb_lineno}: {err1}')
    return str_info


def __check_url(lst_tags, driver, ui, city, keyword):
    lst_firm = []
    index_win = driver.window_handles[0]
    for a_tag in lst_tags:
        second_snap_value = address1 = address2 = url = phone = ''
        try:
            ui.show_message('', '', f'-' * 20)

            second_snap_value = a_tag.get_attribute('aria-label')
            ui.show_message('', '', f'公司名称: {second_snap_value}')

            str_url = a_tag.get_attribute('href')
            # 根据链接 开启新窗口
            driver.execute_script(f"window.open('{str_url}', '_blank')")
            # 切换到新窗口
            new_window = driver.window_handles[-1]
            if new_window != index_win:
                driver.switch_to.window(new_window)
            else:
                ui.show_message('', '', f'未打开新窗口')

            address1 = __get(driver, ui, By.XPATH, "// *[ @ data-item-id='address']", "aria-label").replace('地址:', '').strip()
            ui.show_message('', '', f'公司地址1: {address1}')

            address2 = __get(driver, ui, By.XPATH, "// *[ @ data-item-id='laddress']", "aria-label").replace('地址:', '').strip()
            ui.show_message('', '', f'公司地址2: {address2}')

            url = __get(driver, ui, By.XPATH, "// *[ @ data-tooltip='打开网站']", "href").strip()
            ui.show_message('', '', f'公司网站: {url}')

            phone = __get(driver, ui, By.XPATH, "// *[ @ data-tooltip='复制电话号码']", "aria-label").replace('电话:', '').strip()
            ui.show_message('', '', f'公司电话: {phone}')

            # 关闭新窗口
            if driver.current_window_handle != index_win:
                driver.close()
        except Exception as err:
            logger.error(f'解析失败: {err.__traceback__.tb_lineno}: {err}')
        finally:
            if driver:
                driver.switch_to.window(index_win)
            if any([second_snap_value, address1, address2, url, phone]):
                lst_firm.append({
                    'name': second_snap_value,
                    'address1': address1,
                    'address2': address2,
                    'url': url,
                    'phone': phone,
                    'city': city,
                    'keyword': keyword
                })
    else:
        ui.show_message('', '', f'{city}--{keyword}组合搜索完毕, 共搜索出 {len(lst_firm)} 组信息')
    return lst_firm


def __write_excel(lst_data: list, ui):
    from constant import EMAIL_SEARCH_PATH
    str_file = os.path.join(EMAIL_SEARCH_PATH, f"search_{time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))}.xlsx")
    lst_title = ['城市', '关键字', '公司名称', '公司地址1', '公司地址2', '公司网站', '公司联系电话']
    try:
        wb = Workbook()
        # 默认工作簿
        ws_succ = wb.active
        # 写入表头
        for int_index, str_title in enumerate(lst_title, 1):
            ws_succ.cell(1, int_index).value = str_title
        for int_index, dit_info in enumerate(lst_data, 2):
            ws_succ.cell(int_index, 1).value = dit_info['city']
            ws_succ.cell(int_index, 2).value = dit_info['keyword']
            ws_succ.cell(int_index, 3).value = dit_info['name']
            ws_succ.cell(int_index, 4).value = dit_info['address1']
            ws_succ.cell(int_index, 5).value = dit_info['address2']
            ws_succ.cell(int_index, 6).value = dit_info['url']
            ws_succ.cell(int_index, 7).value = dit_info['phone']
        wb.save(f'{str_file}')
        wb.close()
        ui.show_message('', '', f'邮箱解析结果写入excel成功,保存路径: {str_file}')
    except Exception as err:
        logger.error(f'保存搜索结果失败: {err.__traceback__.tb_lineno}: {err}')
